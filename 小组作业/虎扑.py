from requests_html import HTMLSession
from openpyxl import load_workbook
import re
import time

session = HTMLSession()
url = 'https://m.hupu.com/api/v2/search2?keyword=%E8%B4%A7%E6%8B%89%E6%8B%89%20%E8%87%B4%E6%AD%BB&puid=0&type=posts&topicId=0&page=1'
headers = {
    'cookie': 'smidV2=20221212171431dfb35216a6666c5befc882bd7c6edcbe002f718a4d7cb6ea0; _HUPUSSOID=a6964a5e-1d17-4e76-a6e7-2c2092b7954a; _CLT=b0c2a05996d8b48b354e1fa4ddfc1fef; u=97963809|6JmO5omRSlIwNTM0MTg1Mjkz|4159|6f54b0efd21153334026560244d1e5c6|1eea0b56870b2add|aHVwdV9kNjAzOTVlYjAzYzZiYWEx; us=944d231e865f9c00df1c170f3f966861961b32a14d46948113ae7f73e1a9d62ef504dff65208401ffe468c858e7eb04c500745dbb9a7f8eb53f363b45139543a; ua=167083648; sajssdk_2015_cross_new_user=1; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2218505a14a1846c-007268edab4c7be4-7a575470-3686400-18505a14a191504%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTg1MDVhMTRhMTg0NmMtMDA3MjY4ZWRhYjRjN2JlNC03YTU3NTQ3MC0zNjg2NDAwLTE4NTA1YTE0YTE5MTUwNCJ9%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%22%2C%22value%22%3A%22%22%7D%2C%22%24device_id%22%3A%2218505a14a1846c-007268edab4c7be4-7a575470-3686400-18505a14a191504%22%7D',
    'referer': 'https://m.hupu.com/search',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.46'}
resp = session.get(url=url, headers=headers).json()
# print(resp)
data = resp['data']['result']['data']
# print(data)

os_path = 'C:/Users/1/Desktop/货拉拉.xlsx'
N_J = load_workbook(os_path)
sheet_2 = N_J.active


def filter_tags(htmlstr):
    # 先过滤CDATA
    re_cdata = re.compile('//<!\[CDATA\[[^>]*//\]\]>', re.I)  # 匹配CDATA
    re_script = re.compile('<\s*script[^>]*>[^<]*<\s*/\s*script\s*>', re.I)  # Script
    re_style = re.compile('<\s*style[^>]*>[^<]*<\s*/\s*style\s*>', re.I)  # style
    re_br = re.compile('<br\s*?/?>')  # 处理换行
    re_h = re.compile('</?\w+[^>]*>')  # HTML标签
    re_comment = re.compile('<!--[^>]*-->')  # HTML注释
    s = re_cdata.sub('', htmlstr)  # 去掉CDATA
    s = re_script.sub('', s)  # 去掉SCRIPT
    s = re_style.sub('', s)  # 去掉style
    s = re_br.sub('\n', s)  # 将br转换为换行
    s = re_h.sub('', s)  # 去掉HTML 标签
    s = re_comment.sub('', s)  # 去掉HTML注释
    # 去掉多余的空行
    blank_line = re.compile('\n+')
    s = blank_line.sub('\n', s)
    s = replaceCharEntity(s)  # 替换实体
    return s


def replaceCharEntity(htmlstr):
    CHAR_ENTITIES = {'nbsp': ' ', '160': ' ',
                     'lt': '<', '60': '<',
                     'gt': '>', '62': '>',
                     'amp': '&', '38': '&',
                     'quot': '"', '34': '"', }

    re_charEntity = re.compile(r'&#?(?P<name>\w+);')
    sz = re_charEntity.search(htmlstr)
    while sz:
        entity = sz.group()  # entity全称，如>
        key = sz.group('name')  # 去除&;后entity,如>为gt
        try:
            htmlstr = re_charEntity.sub(CHAR_ENTITIES[key], htmlstr, 1)
            sz = re_charEntity.search(htmlstr)
        except KeyError:
            # 以空串代替
            htmlstr = re_charEntity.sub('', htmlstr, 1)
            sz = re_charEntity.search(htmlstr)
    return htmlstr


row = 2
for i in range(55):
    text = data[i]['title']
    title = filter_tags(text)
    timestamp = int(data[i]['addtime'])
    # time = 1462451334
    # print(time, type(time))

    timeArray = time.localtime(timestamp)
    otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    url = data[i]['link'][21:len(url)]
    # print(url)
    url = 'https://m.hupu.com/bbs/' + url
    print(url)
    html = session.get(url, headers=headers).html
    # content = html.xpath('/html/body/div/div/div[2]/div[1]/div[1]/div/text()')[0]
    # print(content)
    # review = re.findall('<p class="discuss-card__content">()</p>', html.html)
    sheet_2.cell(row, 1, title)
    sheet_2.cell(row, 2, otherStyleTime)
    # sheet_2.cell(row, 3, content)
    sheet_2.cell(row, 6, url)
    row += 1
    N_J.save(os_path)
